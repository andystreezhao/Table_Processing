import pandas as pd

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
full_data = pd.read_excel('resource.xlsx', sheet_name='layout')
full_data = full_data.dropna(axis=1, how='all')  # delete Nan columns
full_data = full_data.dropna(thresh=8)  # reserve rows with 8 or above non-NaN BLOCKS
full_data.to_excel('turn.xlsx', index=False)

import openpyxl

wb = openpyxl.load_workbook(r'turn.xlsx')
sh = wb.active
workbook2 = openpyxl.Workbook()  # get a Workbook
worksheet2 = workbook2.active  # get an active worksheet
kk = sh.cell(1, 1).value[-7:]
for i in range(1, len(full_data) + 2):
    for j in range(1, full_data.columns.size + 1):
        c = sh.cell(i, j).value
        if type(c) == str:
            c = sh.cell(i, j).value.strip().capitalize()
        else:
            c = sh.cell(i, j).value
        # cc = sh.cell(row=i, column=j)  # get the data from the orgin datasheet
        # ee = sh.cell(row=i, column=j).value
        worksheet2.cell(i, j, c)
        # print(c)
        # print(type(c))
for a in range(1, 22):
    worksheet2.insert_cols(3)
    a += 1

#         if type(c) == str:
#             # print(type(c))
#             if j == 1:   #get the postion of the pivot
#                 d = len(c)
#                 if d >= 4:  # fill the pivot row with 'all'
#                     worksheet2.cell(i, 1, c)
#                     ee = 'all'
#                     worksheet2.cell(i, 2, ee)
#                     worksheet2.cell(i, 3, ee)
#                 else:
#                     ee = sh.cell(row=i, column=j).value
#                     worksheet2.cell(i, j, ee)
#
#             if j == 2 or j == 3: #get the maturity and methods columns
#                 d = len(cc.value)
#                 # print(d)
#                 if d < 4:    #get the currency and fill it in the 4th col and fill the 2nd or 3rd with NaN
#                     ee = sh.cell(row=i, column=j).value
#                     ff = sh.cell(row=i, column=4).value
#                     worksheet2.cell(i, 4, ee)
#                     worksheet2.cell(i, j, ff)
#                 else:
#                     ee = sh.cell(row=i, column=j).value
#                     worksheet2.cell(i, j, ee)
#             else:

workbook2.save('result1.xlsx')
# full_data2 = pd.read_excel('result1.xlsx', sheet_name='Sheet')

# workbook2.save('result2.xlsx')
full_data3 = pd.read_excel('result1.xlsx', sheet_name='Sheet')
full_data3.columns = ['code', 'year', 'quarter', 'investment', 'as_li', 'di_comp', 'currency', 'broad', 'standard',
                      'display', 'last_se', 'sb', 'nt', 'chg_no_nt', 'pc', 'xc', 'oc_diff', 'oc', 'oc_total',
                      'chg_period', 'se', 'i', 'a_l', 'co', 's', 'dc', 'cp', 'm', 'cu', 'fld']
# ,'component','sector','transactor','maturity'
full_data3['year'] = kk[-4:]
full_data3['quarter'] = kk[0:2]
full_data3['investment'] = full_data3['code'].str.replace('\n', '')
full_data3['as_li'] = full_data3['code']
full_data3['di_comp'] = full_data3['code']
full_data3['currency'] = full_data3['code']

# df[df.investment.isin([' Direct investment assets'])]
#
# print(full_data3[full_data3.investment.isin(
#     [' Direct investment assets', ' Net international investment position of direct   inventment',
#      ' Direct investment liabilities'])])
# print(full_data3.investment.isin(
#     [' Direct investment assets', ' Net international investment position of direct   inventment',
#      ' Direct investment liabilities']))
s = full_data3.investment.isin(
    ['Direct investment assets', 'Net international investment position of direct   inventment',
     'Direct investment liabilities'])


full_data3['investment'] = s
for b in range(1, len(full_data3)):
    cc = full_data3.iloc[[b-1],[3]]
    # if cc == True:
    print(type(cc))




# full_data3['investment'].loc(full_data3['investment']).isin(
#     [' Direct investment assets', ' Net international investment position of direct   inventment',
#      ' Direct investment liabilities'])
# full_data3.where(full_data3['investment'].isin([' Direct investment assets', ' Net international investment position of direct   inventment',
#       ' Direct investment liabilities']), other = 'empty', inplace = True)
# full_data3.columns=['code','year','quarter','transactor','maturity','component','currency','display','t','m','co','cu','p','fld']
#     # ,'sb','nt','pc','xc','oc','se'
# # full_data3['sb','nt','pc','xc','oc','se']=[[None],]
# full_data3.append[colums['sb','nt','pc','xc','oc','se']]
# print(full_data3.columns.values)
# full_data3.iloc[:, :6] = full_data3.fillna(method='ffill')  # fillna 0-6 col
# full_data3.iloc[:, :7] = full_data3.fillna('all')  # fillna 7th col
full_data3.to_excel('done.xlsx', index=False)  # save without index
